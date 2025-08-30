import os
from datetime import datetime, timedelta
from typing import List, Tuple, Optional, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db import fetch_serials, fetch_serials_by_source, get_serial_stats


def generate_excel(file_path: str, 
                   date_from: Optional[datetime] = None,
                   date_to: Optional[datetime] = None,
                   source: Optional[str] = None,
                   device_type: Optional[str] = None,
                   validation_status: Optional[str] = None) -> str:
    """
    Generate enhanced Excel export with Phase 2.1 fields and formatting.
    
    Args:
        file_path: Path to save the Excel file
        date_from: Start date filter (optional)
        date_to: End date filter (optional)
        source: Source filter (ios/mac/server) (optional)
        device_type: Device type filter (optional)
        validation_status: Validation status filter (valid/invalid) (optional)
    """
    # Fetch data with filters
    rows = fetch_serials()
    
    # Apply filters if provided
    filtered_rows = []
    for row in rows:
        # Unpack row data (assuming order: id, timestamp, serial, device_type, confidence, source, notes, validation_passed, confidence_acceptable)
        if len(row) >= 9:
            row_id, timestamp, serial, device_type_val, confidence, source_val, notes, validation_passed, confidence_acceptable = row[:9]
            
            # Convert timestamp to datetime if it's a string
            if isinstance(timestamp, str):
                try:
                    timestamp_dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                except ValueError:
                    # Skip rows with invalid timestamps
                    continue
            else:
                timestamp_dt = timestamp
            
            # Apply date filter
            if date_from and timestamp_dt < date_from:
                continue
            if date_to and timestamp_dt > date_to:
                continue
            
            # Apply source filter
            if source and source_val != source:
                continue
                
            # Apply device type filter
            if device_type and device_type_val != device_type:
                continue
                
            # Apply validation status filter
            if validation_status:
                if validation_status == "valid" and not validation_passed:
                    continue
                if validation_status == "invalid" and validation_passed:
                    continue
        
            filtered_rows.append(row)
        else:
            # Handle legacy rows without new fields
            filtered_rows.append(row)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create main data sheet
    ws_data = wb.create_sheet("Serial Data")
    
    # Define headers with Phase 2.1 fields
    headers = [
        "ID", "Timestamp (UTC)", "Serial", "Device Type", "Confidence", 
        "Source", "Notes", "Validation Passed", "Confidence Acceptable"
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_idx, row in enumerate(filtered_rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            
            # Format confidence column
            if col_idx == 5 and isinstance(value, (int, float)):  # Confidence column
                cell.number_format = "0.000"
                
            # Format validation columns
            if col_idx in [8, 9]:  # Validation columns
                if value:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws_data.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Add summary statistics
    summary_data = [
        ["Export Information", ""],
        ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")],
        ["Total Records", len(filtered_rows)],
        ["", ""],
        ["Filters Applied", ""],
        ["Date From", date_from.strftime("%Y-%m-%d") if date_from else "All"],
        ["Date To", date_to.strftime("%Y-%m-%d") if date_to else "All"],
        ["Source", source if source else "All"],
        ["Device Type", device_type if device_type else "All"],
        ["Validation Status", validation_status if validation_status else "All"],
        ["", ""],
        ["Data Statistics", ""],
    ]
    
    # Add data statistics
    if filtered_rows:
        # Calculate statistics
        confidences = [row[4] for row in filtered_rows if len(row) > 4 and isinstance(row[4], (int, float))]
        sources = [row[5] for row in filtered_rows if len(row) > 5]
        validation_passed = [row[7] for row in filtered_rows if len(row) > 7]
        confidence_acceptable = [row[8] for row in filtered_rows if len(row) > 8]
        
        if confidences:
            summary_data.extend([
                ["Average Confidence", f"{sum(confidences)/len(confidences):.3f}"],
                ["Min Confidence", f"{min(confidences):.3f}"],
                ["Max Confidence", f"{max(confidences):.3f}"],
            ])
        
        if sources:
            source_counts = {}
            for s in sources:
                source_counts[s] = source_counts.get(s, 0) + 1
            summary_data.append(["", ""])
            summary_data.append(["Records by Source", ""])
            for source_name, count in source_counts.items():
                summary_data.append([f"  {source_name}", count])
        
        if validation_passed:
            valid_count = sum(validation_passed)
            summary_data.extend([
                ["", ""],
                ["Validation Statistics", ""],
                ["Valid Records", valid_count],
                ["Invalid Records", len(validation_passed) - valid_count],
                ["Validation Rate", f"{valid_count/len(validation_passed)*100:.1f}%"],
            ])
    
    # Add summary data to sheet
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)
    
    # Auto-adjust summary column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Create metadata sheet
    ws_metadata = wb.create_sheet("Export Metadata")
    
    metadata = [
        ["Export Metadata", ""],
        ["Generated At", datetime.now().isoformat()],
        ["Total Records Exported", len(filtered_rows)],
        ["Filters Applied", ""],
        ["  Date From", date_from.isoformat() if date_from else "None"],
        ["  Date To", date_to.isoformat() if date_to else "None"],
        ["  Source", source if source else "None"],
        ["  Device Type", device_type if device_type else "None"],
        ["  Validation Status", validation_status if validation_status else "None"],
        ["", ""],
        ["System Information", ""],
        ["Database Records", len(fetch_serials())],
        ["Phase", "2.1"],
        ["Export Version", "1.0"],
    ]
    
    for row_idx, (label, value) in enumerate(metadata, 1):
        ws_metadata.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_metadata.cell(row=row_idx, column=2, value=value)
    
    # Auto-adjust metadata column widths
    for column in ws_metadata.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_metadata.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    
    return file_path


def generate_excel_legacy(file_path: str) -> str:
    """Legacy export function for backward compatibility"""
    rows = fetch_serials()

    wb = Workbook()
    ws = wb.active
    ws.title = "Serials"

    ws.append(["ID", "Timestamp (UTC)", "Serial", "Device Type", "Confidence"]) 
    for row in rows:
        ws.append(list(row))

    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    return file_path
