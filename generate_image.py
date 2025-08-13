import argparse
import csv
import math
import os
import random
import string
from dataclasses import dataclass
from typing import List, Optional, Tuple

import numpy as np
from PIL import Image, ImageDraw, ImageEnhance, ImageFilter, ImageFont

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


# ---------------------------
# Data structures and config
# ---------------------------

@dataclass
class GenConfig:
    width: int = 600
    height: int = 180
    font_size_min: int = 36
    font_size_max: int = 72
    margin_px: int = 16
    invert_prob: float = 0.15
    glare_prob: float = 0.35
    noise_prob: float = 0.6
    blur_sigma_max: float = 1.5
    rotate_deg_max: float = 12.0
    persp_jitter_px: int = 12
    brightness_jitter: Tuple[float, float] = (0.85, 1.15)
    contrast_jitter: Tuple[float, float] = (0.85, 1.25)
    letter_spacing_max: int = 6


# ---------------------------
# Serial generation
# ---------------------------

def generate_serial() -> str:
    """Generates a random 12-character Apple-style serial number."""
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=12))


def _norm_col(s: str) -> str:
    # Lower, replace common O/0 confusion, and strip non-alnum
    s = s.lower().replace("o", "0")
    return "".join(ch for ch in s if ch.isalnum())


def load_serials_from_xlsx(path: str, column: str) -> List[str]:
    if openpyxl is None:
        raise RuntimeError("openpyxl not installed. Install it to read Excel files.")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    want = _norm_col(column)
    header_idx = None

    # First pass: exact normalized match
    for i, h in enumerate(headers):
        if h is None:
            continue
        if _norm_col(str(h)) == want:
            header_idx = i
            break

    # Second pass: substring match (e.g., 'serial' within 'serial n0.')
    if header_idx is None:
        for i, h in enumerate(headers):
            if h is None:
                continue
            if want in _norm_col(str(h)) or "serial" in str(h).lower():
                header_idx = i
                break

    if header_idx is None:
        raise ValueError(f"Column '{column}' not found in Excel header: {headers}")

    values: List[str] = []
    for row in ws.iter_rows(min_row=2):
        v = row[header_idx].value
        if v is None:
            continue
        s = str(v).strip().upper()
        if len(s) == 12:
            values.append(s)
    return values


# ---------------------------
# Rendering helpers
# ---------------------------

def _load_fonts(font_dir: Optional[str]) -> List[ImageFont.FreeTypeFont]:
    candidates: List[str] = []
    if font_dir and os.path.isdir(font_dir):
        for name in os.listdir(font_dir):
            if name.lower().endswith(('.ttf', '.otf')):
                candidates.append(os.path.join(font_dir, name))
    # Common fallbacks if available on system
    for path in [
        "/System/Library/Fonts/SFNS.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        "C:/Windows/Fonts/consola.ttf",
        "C:/Windows/Fonts/seguisym.ttf",
    ]:
        if os.path.exists(path):
            candidates.append(path)

    fonts: List[ImageFont.FreeTypeFont] = []
    for fp in candidates:
        try:
            fonts.append(ImageFont.truetype(fp, size=48))
        except Exception:
            continue
    if not fonts:
        fonts = [ImageFont.load_default()]
    return fonts


def _random_bg_texture(size: Tuple[int, int]) -> Image.Image:
    w, h = size
    base = np.random.normal(loc=245, scale=5, size=(h, w, 1)).astype(np.uint8)
    tint = np.random.normal(loc=245, scale=5, size=(h, w, 1)).astype(np.uint8)
    paper = np.clip(np.concatenate([base, base, tint], axis=2), 0, 255)
    img = Image.fromarray(paper, mode="RGB")
    return img.filter(ImageFilter.GaussianBlur(radius=0.5))


def _draw_text_centered(
    image: Image.Image,
    text: str,
    font: ImageFont.ImageFont,
    fill: Tuple[int, int, int],
    letter_spacing: int,
    margin_px: int,
) -> None:
    draw = ImageDraw.Draw(image)
    x = margin_px
    y_center = image.height // 2
    total_width = 0
    char_sizes: List[Tuple[int, int]] = []
    for ch in text:
        bbox = font.getbbox(ch)
        w = (bbox[2] - bbox[0])
        h = (bbox[3] - bbox[1])
        char_sizes.append((w, h))
        total_width += w + letter_spacing
    total_width -= letter_spacing
    x = max(margin_px, (image.width - total_width) // 2)

    for idx, ch in enumerate(text):
        w, h = char_sizes[idx]
        y = max(margin_px, y_center - h // 2)
        draw.text((x, y), ch, font=font, fill=fill)
        x += w + letter_spacing


def _apply_perspective(img: Image.Image, jitter_px: int) -> Image.Image:
    w, h = img.size
    jitter = lambda v: v + random.randint(-jitter_px, jitter_px)
    src = [(0, 0), (w, 0), (w, h), (0, h)]
    dst = [(jitter(0), jitter(0)), (jitter(w), jitter(0)), (jitter(w), jitter(h)), (jitter(0), jitter(h))]
    coeffs = _find_perspective_coeffs(src, dst)
    return img.transform(img.size, Image.PERSPECTIVE, coeffs, resample=Image.BICUBIC)


def _find_perspective_coeffs(src_pts, dst_pts):
    matrix = []
    for p1, p2 in zip(dst_pts, src_pts):
        matrix.append([p1[0], p1[1], 1, 0, 0, 0, -p2[0] * p1[0], -p2[0] * p1[1]])
        matrix.append([0, 0, 0, p1[0], p1[1], 1, -p2[1] * p1[0], -p2[1] * p1[1]])
    A = np.array(matrix, dtype=float)
    B = np.array([p[0] for p in src_pts] + [p[1] for p in src_pts], dtype=float)
    res = np.linalg.lstsq(A, B, rcond=None)[0]
    return res


def _add_glare(img: Image.Image) -> Image.Image:
    from PIL import ImageDraw as _ImageDraw

    w, h = img.size
    overlay = Image.new("RGBA", img.size, (255, 255, 255, 0))
    draw = _ImageDraw.Draw(overlay)
    band_h = random.randint(max(6, h // 15), max(12, h // 8))
    y = random.randint(0, max(0, h - band_h))
    alpha_top = random.randint(20, 60)
    alpha_bottom = 0
    for i in range(band_h):
        a = int(alpha_top + (alpha_bottom - alpha_top) * (i / max(1, band_h - 1)))
        draw.line([(0, y + i), (w, y + i)], fill=(255, 255, 255, a), width=1)
    combined = Image.alpha_composite(img.convert("RGBA"), overlay)
    return combined.convert("RGB")


def _apply_distortions(img: Image.Image, cfg: GenConfig) -> Image.Image:
    deg = random.uniform(-cfg.rotate_deg_max, cfg.rotate_deg_max)
    img = img.rotate(deg, expand=True, resample=Image.BICUBIC, fillcolor=(255, 255, 255))
    img = _apply_perspective(img, cfg.persp_jitter_px)
    img = ImageEnhance.Brightness(img).enhance(random.uniform(*cfg.brightness_jitter))
    img = ImageEnhance.Contrast(img).enhance(random.uniform(*cfg.contrast_jitter))
    if random.random() < cfg.glare_prob:
        img = _add_glare(img)
    if random.random() < cfg.noise_prob:
        np_img = np.array(img)
        noise = np.random.normal(0, 8, np_img.shape).astype(np.int16)
        np_img = np.clip(np_img.astype(np.int16) + noise, 0, 255).astype(np.uint8)
        img = Image.fromarray(np_img)
    blur_sigma = random.uniform(0, cfg.blur_sigma_max)
    if blur_sigma > 0.05:
        img = img.filter(ImageFilter.GaussianBlur(radius=blur_sigma))
    return img


# ---------------------------
# Main generator
# ---------------------------

def create_synthetic_image(
    serial: str,
    file_path: str,
    font: ImageFont.ImageFont,
    cfg: GenConfig,
    invert_colors: bool = False,
    letter_spacing: int = 0,
) -> None:
    bg = _random_bg_texture((cfg.width, cfg.height))
    text_color = (0, 0, 0)
    bg_color = (255, 255, 255)
    if invert_colors:
        text_color, bg_color = bg_color, text_color
        bg = Image.new("RGB", (cfg.width, cfg.height), bg_color)
    _draw_text_centered(bg, serial, font, text_color, letter_spacing, cfg.margin_px)
    img = _apply_distortions(bg, cfg)
    img = img.resize((cfg.width, cfg.height), resample=Image.BICUBIC)
    img.save(file_path)


def generate_dataset(
    out_dir: str,
    num_images: int,
    font_dir: Optional[str],
    seed: Optional[int],
    formats: Tuple[str, ...] = ("png",),
    from_xlsx: Optional[str] = None,
    xlsx_column: str = "serial",
    mask_last: int = 0,
) -> None:
    if seed is not None:
        random.seed(seed)
        np.random.seed(seed)

    os.makedirs(out_dir, exist_ok=True)
    labels_path = os.path.join(out_dir, "labels.csv")

    fonts = _load_fonts(font_dir)
    cfg = GenConfig()

    serials: List[str]
    if from_xlsx:
        serials = load_serials_from_xlsx(from_xlsx, xlsx_column)
        if num_images > 0:
            serials = serials[:num_images]
    else:
        serials = [generate_serial() for _ in range(num_images)]

    def _mask(s: str) -> str:
        if mask_last <= 0 or mask_last >= len(s):
            return s
        return s[:-mask_last] + ("X" * mask_last)

    with open(labels_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["filename", "serial", "masked"])
        for i, serial in enumerate(serials, start=1):
            font_size = random.randint(cfg.font_size_min, cfg.font_size_max)
            font = random.choice(fonts)
            try:
                if isinstance(font, ImageFont.FreeTypeFont):
                    font = ImageFont.truetype(font.path, font_size)
    except Exception:
                pass

            invert = random.random() < cfg.invert_prob
            spacing = random.randint(0, cfg.letter_spacing_max)
            label = []
            if invert:
                label.append("inv")
            if spacing:
                label.append(f"sp{spacing}")
            base_name = f"synthetic_{i:04d}_{'_'.join(label) if label else 'clean'}_{_mask(serial)}"

            for ext in formats:
                filename = f"{base_name}.{ext}"
                path = os.path.join(out_dir, filename)
                create_synthetic_image(
                    serial=serial,
                    file_path=path,
                    font=font,
                    cfg=cfg,
                    invert_colors=invert,
                    letter_spacing=spacing,
                )
                writer.writerow([filename, serial, _mask(serial)])


# ---------------------------
# CLI
# ---------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Generate synthetic Apple serial images")
    parser.add_argument("--output-dir", default="synthetic_test_images", help="Output directory (private)")
    parser.add_argument("--num", type=int, default=50, help="Number of images to generate (ignored if --from-xlsx is used)")
    parser.add_argument("--font-dir", default=None, help="Optional directory containing .ttf/.otf fonts")
    parser.add_argument("--formats", default="png", help="Comma-separated list of formats (e.g., png,jpg)")
    parser.add_argument("--seed", type=int, default=None, help="Random seed for reproducibility")
    parser.add_argument("--from-xlsx", default=None, help="Path to Excel sheet with real serials")
    parser.add_argument("--xlsx-column", default="serial", help="Column name that contains serials")
    parser.add_argument("--mask-last", type=int, default=4, help="Mask last N chars in filenames/labels")
    args = parser.parse_args()

    out_dir = args.output_dir
    if not out_dir.lower().startswith("private"):
        out_dir = os.path.join("private", out_dir)

    formats = tuple([fmt.strip().lower() for fmt in args.formats.split(",") if fmt.strip()])
    generate_dataset(out_dir, args.num, args.font_dir, args.seed, formats, args.from_xlsx, args.xlsx_column, args.mask_last)
    print(f"Generated images in '{out_dir}' with labels.csv (masked last {args.mask_last} chars)")


if __name__ == "__main__":
    main()

